#!/usr/bin/env python3
"""
Consolidate all MCAT and MedMCQA evaluation results into a single Excel workbook.

Sheets:
  1. Summary           - Overall accuracy comparison across all models and benchmarks
  2. MCAT_Finetuned    - MCAT results: Qwen3-14B LoRA + Qwen3.5-9B GGUF (fine-tuned on MedMCQA)
  3. MCAT_Local_GGUF   - MCAT results: local GGUF zero-shot direct prompt
  4. MCAT_Local_CoT    - MCAT results: local GGUF zero-shot with Chain-of-Thought
  5. MedMCQA_Eval      - MedMCQA dev-set accuracy: baselines + Gemma LoRA
  6. MedMCQA_PerSubject- Per-subject accuracy breakdown for MedMCQA models
  7. DeBERTa_Training  - DeBERTa-v3-large cross-encoder training metrics (epoch-by-epoch)
"""

import json
import os
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
TEST_SETS = BASE / "test sets"
MEDMCQA = BASE / "medmcqa"
MED = BASE / "med"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def autofit_columns(ws, min_width=10, max_width=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ALT_FILL    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def style_sheet(ws):
    for i, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if i == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            elif i % 2 == 1:
                cell.fill = ALT_FILL
    autofit_columns(ws)


def write_df(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    style_sheet(ws)
    return ws


# ---------------------------------------------------------------------------
# 1. MCAT Finetuned (DGX8 results pulled locally)
# ---------------------------------------------------------------------------

def load_mcat_finetuned():
    summary_split = TEST_SETS / "results_finetuned" / "summary_by_split.csv"
    summary_agg   = TEST_SETS / "results_finetuned" / "summary_aggregated.csv"
    if summary_split.exists():
        df_split = pd.read_csv(summary_split)
    else:
        # build from per_model CSVs
        parts = []
        for f in sorted((TEST_SETS / "results_finetuned" / "per_model").glob("*.csv")):
            parts.append(pd.read_csv(f))
        df_split = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

    df_agg = pd.read_csv(summary_agg) if summary_agg.exists() else pd.DataFrame()
    return df_split, df_agg


# ---------------------------------------------------------------------------
# 2. MCAT Local GGUF (direct prompt, no CoT)
# ---------------------------------------------------------------------------

def load_mcat_local(results_dir):
    summary_split = results_dir / "summary_by_split.csv"
    summary_agg   = results_dir / "summary_aggregated.csv"
    df_split = pd.read_csv(summary_split) if summary_split.exists() else pd.DataFrame()
    df_agg   = pd.read_csv(summary_agg)   if summary_agg.exists()   else pd.DataFrame()
    return df_split, df_agg


# ---------------------------------------------------------------------------
# 3. MedMCQA evaluation (metrics JSON files)
# ---------------------------------------------------------------------------

def load_medmcqa_eval():
    results_dir = MEDMCQA / "results"
    rows = []
    per_subject_rows = []

    for model_dir in sorted(results_dir.iterdir()):
        if not model_dir.is_dir():
            continue
        metrics_file = model_dir / f"{model_dir.name}_metrics.json"
        if not metrics_file.exists():
            # try any *_metrics.json
            candidates = list(model_dir.glob("*_metrics.json"))
            if not candidates:
                continue
            metrics_file = candidates[0]

        with open(metrics_file) as f:
            m = json.load(f)

        rows.append({
            "model": model_dir.name,
            "overall_accuracy": round(m.get("overall_accuracy", 0) * 100, 2),
            "macro_avg_accuracy": round(m.get("macro_averaged_accuracy", 0) * 100, 2),
            "total_samples": m.get("total_samples", ""),
            "correct": m.get("correct", ""),
        })

        for subject, stats in m.get("per_subject_accuracy", {}).items():
            per_subject_rows.append({
                "model": model_dir.name,
                "subject": subject,
                "accuracy_%": round(stats["accuracy"] * 100, 2),
                "n": stats["n"],
            })

    df_summary = pd.DataFrame(rows).sort_values("overall_accuracy", ascending=False).reset_index(drop=True)
    df_subject = pd.DataFrame(per_subject_rows).sort_values(["subject", "model"]).reset_index(drop=True)
    return df_summary, df_subject


# ---------------------------------------------------------------------------
# 4. DeBERTa training metrics
# ---------------------------------------------------------------------------

def load_deberta():
    p = MED / "results" / "deberta_metrics.csv"
    if not p.exists():
        return pd.DataFrame()
    df = pd.read_csv(p)
    # Round floats
    float_cols = df.select_dtypes(include="float").columns
    df[float_cols] = df[float_cols].round(4)
    return df


# ---------------------------------------------------------------------------
# 5. Summary sheet — best accuracy per model × benchmark
# ---------------------------------------------------------------------------

def build_summary(ft_agg, local_agg, local_cot_agg, medmcqa_df):
    rows = []

    # MCAT finetuned
    for _, r in ft_agg.iterrows():
        rows.append({
            "Model": r.get("model_display_name", r.get("model_key", "")),
            "Benchmark": "MCAT (7 test sets)",
            "Evaluation type": "Fine-tuned (MedMCQA LoRA/GGUF)",
            "Accuracy %": round(r.get("mean_accuracy", 0) * 100, 2),
            "Std %": round(r.get("std_accuracy", 0) * 100, 2),
            "Mean F1": round(r.get("mean_macro_f1", 0), 4),
            "N questions": r.get("total_questions", ""),
        })

    # MCAT local GGUF
    for _, r in local_agg.iterrows():
        rows.append({
            "Model": r.get("model_display_name", r.get("model_key", "")),
            "Benchmark": "MCAT (7 test sets)",
            "Evaluation type": "Zero-shot GGUF (direct prompt)",
            "Accuracy %": round(r.get("mean_accuracy", 0) * 100, 2),
            "Std %": round(r.get("std_accuracy", 0) * 100, 2),
            "Mean F1": round(r.get("mean_macro_f1", 0), 4),
            "N questions": r.get("total_questions", ""),
        })

    # MCAT local CoT
    for _, r in local_cot_agg.iterrows():
        rows.append({
            "Model": r.get("model_display_name", r.get("model_key", "")),
            "Benchmark": "MCAT (7 test sets)",
            "Evaluation type": "Zero-shot GGUF (Chain-of-Thought)",
            "Accuracy %": round(r.get("mean_accuracy", 0) * 100, 2),
            "Std %": round(r.get("std_accuracy", 0) * 100, 2),
            "Mean F1": round(r.get("mean_macro_f1", 0), 4),
            "N questions": r.get("total_questions", ""),
        })

    # MedMCQA
    for _, r in medmcqa_df.iterrows():
        rows.append({
            "Model": r["model"],
            "Benchmark": "MedMCQA (dev set, 4183 q)",
            "Evaluation type": "Fine-tuned or zero-shot (MedMCQA)",
            "Accuracy %": r["overall_accuracy"],
            "Std %": "",
            "Mean F1": "",
            "N questions": r["total_samples"],
        })

    df = pd.DataFrame(rows, columns=["Model", "Benchmark", "Evaluation type",
                                      "Accuracy %", "Std %", "Mean F1", "N questions"])
    return df.sort_values(["Benchmark", "Accuracy %"], ascending=[True, False]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    import openpyxl

    print("Loading data...")

    ft_split, ft_agg = load_mcat_finetuned()
    local_split, local_agg = load_mcat_local(TEST_SETS / "results_local_llm")
    cot_split, cot_agg = load_mcat_local(TEST_SETS / "results_local_llm_cot")
    medmcqa_df, medmcqa_subject_df = load_medmcqa_eval()
    deberta_df = load_deberta()

    summary_df = build_summary(ft_agg, local_agg, cot_agg, medmcqa_df)
    speed_df = build_speed_summary()

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Writing sheets...")

    write_df(wb, "Summary", summary_df)

    if not ft_split.empty:
        write_df(wb, "MCAT_Finetuned_BySplit", ft_split)
    if not ft_agg.empty:
        write_df(wb, "MCAT_Finetuned_Agg", ft_agg)

    if not local_split.empty:
        write_df(wb, "MCAT_Local_GGUF_BySplit", local_split)
    if not local_agg.empty:
        write_df(wb, "MCAT_Local_GGUF_Agg", local_agg)

    if not cot_split.empty:
        write_df(wb, "MCAT_Local_CoT_BySplit", cot_split)
    if not cot_agg.empty:
        write_df(wb, "MCAT_Local_CoT_Agg", cot_agg)

    if not medmcqa_df.empty:
        write_df(wb, "MedMCQA_Eval", medmcqa_df)
    if not medmcqa_subject_df.empty:
        write_df(wb, "MedMCQA_PerSubject", medmcqa_subject_df)

    if not deberta_df.empty:
        write_df(wb, "DeBERTa_Training", deberta_df)

    if not speed_df.empty:
        write_df(wb, "Speed_Summary", speed_df)

    out = BASE / "results_consolidated.xlsx"
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")

    # Also export flat CSVs for sharing
    print("\nExporting flat CSVs...")
    csv_out = BASE / "consolidated_csvs"

    # Build all_predictions for MCAT finetuned from per_model CSVs (already has all cols)
    ft_all = pd.read_csv(TEST_SETS / "results_finetuned" / "all_predictions.csv") \
        if (TEST_SETS / "results_finetuned" / "all_predictions.csv").exists() else pd.DataFrame()
    local_all = pd.read_csv(TEST_SETS / "results_local_llm" / "all_predictions.csv") \
        if (TEST_SETS / "results_local_llm" / "all_predictions.csv").exists() else pd.DataFrame()
    cot_all = pd.read_csv(TEST_SETS / "results_local_llm_cot" / "all_predictions.csv") \
        if (TEST_SETS / "results_local_llm_cot" / "all_predictions.csv").exists() else pd.DataFrame()

    # MedMCQA predictions — combine all three model prediction CSVs
    medmcqa_preds = []
    for model_dir in sorted((MEDMCQA / "results").iterdir()):
        for f in model_dir.glob("*_predictions.csv"):
            df = pd.read_csv(f)
            df.insert(0, "model", model_dir.name)
            medmcqa_preds.append(df)
    medmcqa_all_preds = pd.concat(medmcqa_preds, ignore_index=True) if medmcqa_preds else pd.DataFrame()

    export_csvs({
        "01_summary":                    summary_df,
        "02_mcat_finetuned_by_split":    ft_split,
        "02_mcat_finetuned_aggregated":  ft_agg,
        "02_mcat_finetuned_all_predictions": ft_all,
        "03_mcat_local_gguf_by_split":   local_split,
        "03_mcat_local_gguf_aggregated": local_agg,
        "03_mcat_local_gguf_all_predictions": local_all,
        "04_mcat_local_cot_by_split":    cot_split,
        "04_mcat_local_cot_aggregated":  cot_agg,
        "04_mcat_local_cot_all_predictions": cot_all,
        "05_medmcqa_eval_summary":       medmcqa_df,
        "05_medmcqa_per_subject":        medmcqa_subject_df,
        "05_medmcqa_all_predictions":    medmcqa_all_preds,
        "06_deberta_training":           deberta_df,
        "07_speed_summary":              speed_df,
    }, csv_out)
    print(f"\nAll CSVs in: {csv_out}")


# ---------------------------------------------------------------------------
# Speed summary — training + inference, from logs + prediction latency columns
# ---------------------------------------------------------------------------

def build_speed_summary():
    rows = []

    # ── Training speed (from log grep) ──────────────────────────────────────
    # Source: DGX8 medmcqa/logs/  (runtime in seconds from trainer JSON)
    training = [
        # (model, hardware, dtype, steps, runtime_s, samples_per_s, steps_per_s, train_loss, notes)
        ("Qwen3-14B LoRA",    "DGX8 GB10 (121GB)", "BF16",   500,   99550, 0.161, 0.010, 1.147, "archive run; ~27.6 hrs; ~65-70 s/step"),
        ("Qwen3.5-9B LoRA",   "DGX8 GB10 (121GB)", "BF16",  1000,   63190, 0.253, 0.016, 1.120, "early stop at 1000 steps; ~17.5 hrs; ~45-55 s/step"),
        ("Gemma-3-4B LoRA",   "DGX8 GB10 (121GB)", "BF16",  2000,   51090, 0.626, 0.039, 1.325, "2000-step run; ~14.2 hrs; ~25 s/step"),
        ("Gemma-3-4B LoRA",   "MBP M-series (MPS)","BF16",  None,   None,  None,  None, None, "~29-31 s/step on MPS; impractical for full run"),
    ]
    for t in training:
        model, hw, dtype, steps, rt, sps, stps, loss, notes = t
        rows.append({
            "Category": "Training",
            "Model": model,
            "Hardware": hw,
            "Precision": dtype,
            "Steps / Epochs": steps if steps != "—" else "—",
            "Runtime (s)": rt if rt != "—" else "—",
            "Runtime (hrs)": round(rt / 3600, 2) if isinstance(rt, (int, float)) else "—",
            "Samples/sec": sps if sps != "—" else "—",
            "Steps/sec": stps if stps != "—" else "—",
            "Train loss": loss if loss != "—" else "—",
            "Mean latency/q (s)": "—",
            "Median latency/q (s)": "—",
            "P90 latency/q (s)": "—",
            "Mean tokens/sec": "—",
            "Total eval time (min)": "—",
            "N questions": "—",
            "Notes": notes,
        })

    # ── Inference speed (from all_predictions latency_s columns) ────────────
    inference_sources = [
        (TEST_SETS / "results_finetuned"     / "all_predictions.csv", "MCAT Inference (DGX8 BF16/GGUF, fine-tuned)"),
        (TEST_SETS / "results_local_llm"     / "all_predictions.csv", "MCAT Inference (Local MBP, GGUF zero-shot)"),
        (TEST_SETS / "results_local_llm_cot" / "all_predictions.csv", "MCAT Inference (Local MBP, GGUF CoT)"),
    ]
    for path, category in inference_sources:
        if not path.exists():
            continue
        df = pd.read_csv(path)
        for model_name, grp in df.groupby("model_display_name"):
            lat = grp["latency_s"]
            tok_s = grp["tokens_per_sec"] if "tokens_per_sec" in grp else None
            rows.append({
                "Category": category,
                "Model": model_name,
                "Hardware": "DGX8 GB10" if "DGX8" in category else "MacBook Pro (MPS/CPU)",
                "Precision": "BF16" if "LoRA" in model_name else "Q4_K_M" if "Q4_K_M" in model_name else "q4_0",
                "Steps / Epochs": "—",
                "Runtime (s)": "—",
                "Runtime (hrs)": "—",
                "Samples/sec": "—",
                "Steps/sec": "—",
                "Train loss": "—",
                "Mean latency/q (s)": round(lat.mean(), 2),
                "Median latency/q (s)": round(lat.median(), 2),
                "P90 latency/q (s)": round(lat.quantile(0.9), 2),
                "Mean tokens/sec": round(tok_s.mean(), 2) if tok_s is not None else "—",
                "Total eval time (min)": round(lat.sum() / 60, 1),
                "N questions": len(grp),
                "Notes": "",
            })

    # MedMCQA baseline inference — from log: eval tqdm completed 4183 questions
    # 14B: log shows tqdm running through 523 batches (bs=8) → ~4183q
    # Times derived from log wall-clock (eval started/finished same session)
    medmcqa_inf = [
        ("Qwen3-14B (zero-shot)", "DGX8 GB10", "BF16", 4183, "~3.5-4 hrs", "~3-4 s/q est.", "MedMCQA baseline eval"),
        ("Qwen3.5-9B (zero-shot)", "DGX8 GB10", "BF16", 4183, "~3-3.5 hrs", "~2.5-3 s/q est.", "MedMCQA baseline eval"),
        ("Gemma-3-4B LoRA",        "DGX8 GB10", "BF16", 4183, "~2-2.5 hrs", "~1.7-2 s/q est.", "MedMCQA lora eval"),
    ]
    for model, hw, dtype, nq, total, lat, notes in medmcqa_inf:
        rows.append({
            "Category": "MedMCQA Inference (DGX8)",
            "Model": model,
            "Hardware": hw,
            "Precision": dtype,
            "Steps / Epochs": "—",
            "Runtime (s)": "—",
            "Runtime (hrs)": "—",
            "Samples/sec": "—",
            "Steps/sec": "—",
            "Train loss": "—",
            "Mean latency/q (s)": lat,
            "Median latency/q (s)": "—",
            "P90 latency/q (s)": "—",
            "Mean tokens/sec": "—",
            "Total eval time (min)": total,
            "N questions": nq,
            "Notes": notes,
        })

    cols = ["Category", "Model", "Hardware", "Precision",
            "Steps / Epochs", "Runtime (s)", "Runtime (hrs)", "Samples/sec", "Steps/sec", "Train loss",
            "Mean latency/q (s)", "Median latency/q (s)", "P90 latency/q (s)", "Mean tokens/sec",
            "Total eval time (min)", "N questions", "Notes"]
    return pd.DataFrame(rows, columns=cols)


# ---------------------------------------------------------------------------
# Export flat CSVs for sharing
# ---------------------------------------------------------------------------

def export_csvs(dfs: dict, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    for name, df in dfs.items():
        if df is not None and not df.empty:
            p = out_dir / f"{name}.csv"
            df.to_csv(p, index=False)
            print(f"  {p.name}  ({len(df)} rows)")


if __name__ == "__main__":
    main()
